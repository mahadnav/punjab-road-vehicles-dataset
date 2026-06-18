import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from copy import copy

INPUT_FILE = r"c:\Users\Acer\Desktop\Punjab EI\Punjab Transport.xlsx"
RAW_SHEET  = "Raw"      # raw data (6155 rows)
OUT_SHEET  = "Cleaned"

df = pd.read_excel(INPUT_FILE, sheet_name=RAW_SHEET)
df.columns = ["Year", "Province", "Division", "District", "VehicleType", "Count"]
df["Year"]  = df["Year"].astype(int)
df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0)

print(f"Loaded {len(df)} rows | Years: {sorted(df['Year'].unique())}")

dims = ["District", "VehicleType"]   # keys for every fix lookup
changes = []                          # audit log

# Helper
def get(pivot, dist, vt, year):
    try:
        return pivot.loc[(dist, vt), year]
    except KeyError:
        return None

def linear_interp(lo_yr, lo_val, hi_yr, hi_val, target_yr):
    """Simple linear interpolation between two known points."""
    span = hi_yr - lo_yr
    if span == 0:
        return lo_val
    return lo_val + (hi_val - lo_val) * (target_yr - lo_yr) / span

def two_step_interp(lo_yr, lo_val, hi_yr, hi_val, yr1, yr2):
    """Spread evenly across two corrupted interior years."""
    v1 = linear_interp(lo_yr, lo_val, hi_yr, hi_val, yr1)
    v2 = linear_interp(lo_yr, lo_val, hi_yr, hi_val, yr2)
    return round(v1), round(v2)

# Build a pivot for easy lookup: index=(District, VehicleType), cols=Year
pivot = df.pivot_table(index=dims, columns="Year", values="Count", aggfunc="sum")
districts  = df["District"].unique()
vtypes     = df["VehicleType"].unique()
all_years  = sorted(df["Year"].unique())

# Apply fixes, building corrected_values dict
# corrected_values[(dist, vt, year)] = new_count
corrected = {}

for dist in districts:
    for vt in vtypes:

        vals = {}
        for y in all_years:
            v = get(pivot, dist, vt, y)
            vals[y] = float(v) if (v is not None and not np.isnan(v)) else None

        # FIX 1 & 2: 2018 flat copy of 2017
        # Applies to: Auto Rickshaws, Motor Cycles/Scooters, Pickups,
        #             Taxis, Tractors, Trucks (but NOT MinisBuses, Motor Cars)
        # Detect: 2018 == 2017 AND both are non-zero
        v17 = vals.get(2017)
        v18 = vals.get(2018)
        v19 = vals.get(2019)
        if (v17 is not None and v18 is not None and v19 is not None
                and v17 == v18 and v17 > 0):
            new18 = round(linear_interp(2017, v17, 2019, v19, 2018))
            corrected[(dist, vt, 2018)] = new18
            changes.append({
                "District": dist, "VehicleType": vt, "Year": 2018,
                "OldValue": int(v18), "NewValue": new18,
                "Rule": "Fix1: 2018 flat copy of 2017 → linear interp(2017,2019)"
            })

        # FIX 3: 2022 missing entirely
        v21 = vals.get(2021)
        v23 = vals.get(2023)
        if v21 is not None and v23 is not None and v21 > 0:
            new22 = round(linear_interp(2021, v21, 2023, v23, 2022))
            corrected[(dist, vt, 2022)] = new22
            changes.append({
                "District": dist, "VehicleType": vt, "Year": 2022,
                "OldValue": None, "NewValue": new22,
                "Rule": "Fix3: 2022 missing → linear interp(2021,2023)"
            })

        # FIX 4: Trucks 2015 spike
        if vt == "Trucks":
            v14 = vals.get(2014)
            v15 = vals.get(2015)
            v16 = vals.get(2016)
            if (v14 is not None and v15 is not None and v16 is not None
                    and v14 > 0 and v16 > 0):
                ratio = v15 / ((v14 + v16) / 2) if (v14 + v16) > 0 else 1
                if ratio > 1.5:   # spike threshold: 50% above average of neighbours
                    new15 = round(linear_interp(2014, v14, 2016, v16, 2015))
                    corrected[(dist, vt, 2015)] = new15
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2015,
                        "OldValue": int(v15), "NewValue": new15,
                        "Rule": "Fix4: Trucks 2015 spike → linear interp(2014,2016)"
                    })

        # FIX 5: Mini Buses 2017 spike
        if vt == "Mini Buses/Buses/Flying/Luxury Coaches":
            v16 = vals.get(2016)
            v17 = vals.get(2017)
            v18 = vals.get(2018)
            if (v16 is not None and v17 is not None and v18 is not None
                    and v16 > 0 and v18 > 0):
                ratio = v17 / ((v16 + v18) / 2) if (v16 + v18) > 0 else 1
                if ratio > 1.5:
                    new17 = round(linear_interp(2016, v16, 2018, v18, 2017))
                    corrected[(dist, vt, 2017)] = new17
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2017,
                        "OldValue": int(v17), "NewValue": new17,
                        "Rule": "Fix5: MiniBuses 2017 spike → linear interp(2016,2018)"
                    })

        # FIX 6: Tractors 2017 drop + 2018 flat copy
        if vt == "Tractors":
            v16 = vals.get(2016)
            v17 = vals.get(2017)
            v18 = vals.get(2018)
            v19 = vals.get(2019)
            if (v16 is not None and v19 is not None and v16 > 0 and v19 > 0
                    and v17 is not None and v17 > 0):
                drop_ratio = v17 / v16
                flat_copy  = (v17 == v18)
                if drop_ratio < 0.75 and flat_copy:
                    # Both years corrupted → spread linearly 2016→2019
                    new17, new18 = two_step_interp(2016, v16, 2019, v19, 2017, 2018)
                    corrected[(dist, vt, 2017)] = new17
                    corrected[(dist, vt, 2018)] = new18
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2017,
                        "OldValue": int(v17), "NewValue": new17,
                        "Rule": "Fix6: Tractors 2017 drop + 2018 flat → 2-step interp(2016,2019)"
                    })
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2018,
                        "OldValue": int(v18), "NewValue": new18,
                        "Rule": "Fix6: Tractors 2017 drop + 2018 flat → 2-step interp(2016,2019)"
                    })

        # FIX 7: Other Vehicles boundary shifts
        # 2011 spike (if 2011 > 3× average of 2010 and 2012)
        # 2017 drop  (if 2017 < 0.4× average of 2016 and 2018)
        if vt == "Other Vehicles":
            v10 = vals.get(2010)
            v11 = vals.get(2011)
            v12 = vals.get(2012)
            if (v10 is not None and v11 is not None and v12 is not None
                    and v10 > 0 and v12 > 0):
                if v11 > 1.5 * v10 and v11 > 2 * v10:
                    new11 = round(linear_interp(2010, v10, 2012, v12, 2011))
                    corrected[(dist, vt, 2011)] = new11
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2011,
                        "OldValue": int(v11), "NewValue": new11,
                        "Rule": "Fix7a: OtherVehicles 2011 spike → linear interp(2010,2012)"
                    })

            v16 = vals.get(2016)
            v17 = vals.get(2017)
            v18 = vals.get(2018)
            if (v16 is not None and v17 is not None and v18 is not None
                    and v16 > 0):
                if v17 < 0.5 * v16 and v17 > 0:
                    anchor_hi = v18 if v18 and v18 > v17 else None
                    if anchor_hi:
                        new17 = round(linear_interp(2016, v16, 2018, anchor_hi, 2017))
                    else:
                        new17 = round((v16 + v17) / 2)  # fallback
                    corrected[(dist, vt, 2017)] = new17
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2017,
                        "OldValue": int(v17), "NewValue": new17,
                        "Rule": "Fix7b: OtherVehicles 2017 drop → linear interp(2016,2018)"
                    })

        # FIX 8: Pickups 2012 spike and 2015 dip
        if vt == "Pickups/Delivery Vans":
            v11 = vals.get(2011)
            v12 = vals.get(2012)
            v13 = vals.get(2013)
            if (v11 is not None and v12 is not None and v13 is not None
                    and v11 > 0 and v13 > 0):
                if v12 > 1.4 * ((v11 + v13) / 2):
                    new12 = round(linear_interp(2011, v11, 2013, v13, 2012))
                    corrected[(dist, vt, 2012)] = new12
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2012,
                        "OldValue": int(v12), "NewValue": new12,
                        "Rule": "Fix8a: Pickups 2012 spike → linear interp(2011,2013)"
                    })

            v14 = vals.get(2014)
            v15 = vals.get(2015)
            v16 = vals.get(2016)
            if (v14 is not None and v15 is not None and v16 is not None
                    and v14 > 0 and v16 > 0):
                if v15 < 0.8 * ((v14 + v16) / 2):
                    new15 = round(linear_interp(2014, v14, 2016, v16, 2015))
                    corrected[(dist, vt, 2015)] = new15
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2015,
                        "OldValue": int(v15), "NewValue": new15,
                        "Rule": "Fix8b: Pickups 2015 dip → linear interp(2014,2016)"
                    })

        # FIX 9: Taxis 2004 drop and 2012 spike
        if vt == "Taxis":
            v03 = vals.get(2003)
            v04 = vals.get(2004)
            v05 = vals.get(2005)
            if (v03 is not None and v04 is not None and v05 is not None
                    and v03 > 0 and v05 > 0):
                if v04 < 0.75 * ((v03 + v05) / 2):
                    new04 = round(linear_interp(2003, v03, 2005, v05, 2004))
                    corrected[(dist, vt, 2004)] = new04
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2004,
                        "OldValue": int(v04), "NewValue": new04,
                        "Rule": "Fix9a: Taxis 2004 drop → linear interp(2003,2005)"
                    })

            v11 = vals.get(2011)
            v12 = vals.get(2012)
            v13 = vals.get(2013)
            if (v11 is not None and v12 is not None and v13 is not None
                    and v11 > 0 and v13 > 0):
                if v12 > 1.3 * v13 and v12 > 1.5 * v11:
                    new12 = round(linear_interp(2011, v11, 2013, v13, 2012))
                    corrected[(dist, vt, 2012)] = new12
                    changes.append({
                        "District": dist, "VehicleType": vt, "Year": 2012,
                        "OldValue": int(v12), "NewValue": new12,
                        "Rule": "Fix9b: Taxis 2012 spike → linear interp(2011,2013)"
                    })

print(f"\nTotal corrections: {len(changes)}")

# Build cleaned dataframe
df_clean = df.copy()

# Apply corrections to existing rows
for idx, row in df_clean.iterrows():
    key = (row["District"], row["VehicleType"], row["Year"])
    if key in corrected:
        df_clean.at[idx, "Count"] = corrected[key]

# Add 2022 rows (new rows, not present in original)
new_rows = []
for (dist, vt, yr), val in corrected.items():
    if yr == 2022:
        # find province and division from existing rows
        ref = df[(df["District"] == dist) & (df["VehicleType"] == vt)].iloc[0]
        new_rows.append({
            "Year": 2022,
            "Province": ref["Province"],
            "Division": ref["Division"],
            "District": dist,
            "VehicleType": vt,
            "Count": val
        })

if new_rows:
    df_new = pd.DataFrame(new_rows)
    df_clean = pd.concat([df_clean, df_new], ignore_index=True)
    df_clean = df_clean.sort_values(["Year", "District", "VehicleType"]).reset_index(drop=True)

print(f"Cleaned dataset rows: {len(df_clean)}")
print(f"Years in cleaned data: {sorted(df_clean['Year'].unique())}")

yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
changed_keys = set(corrected.keys())

with pd.ExcelWriter(INPUT_FILE, engine="openpyxl", mode="a",
                    if_sheet_exists="replace") as writer:
    df_clean.to_excel(writer, sheet_name=OUT_SHEET, index=False)

    ws_clean = writer.sheets[OUT_SHEET]
    # Bold header
    for cell in ws_clean[1]:
        cell.font = Font(bold=True)
    # Highlight corrected Count cells (column 6)
    for r, row in enumerate(df_clean.itertuples(index=False), start=2):
        key = (row.District, row.VehicleType, int(row.Year))
        if key in changed_keys:
            ws_clean.cell(r, 6).fill = yellow

print(f"\nSaved '{OUT_SHEET}' sheet to {INPUT_FILE}")

changes_df = pd.DataFrame(changes)
print("\n=== AUDIT LOG SUMMARY ===")
print(changes_df.groupby("Rule").size().to_string())
print(f"\nTotal cells corrected: {len(changes_df)}")

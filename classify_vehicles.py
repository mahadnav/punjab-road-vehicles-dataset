import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE = r"c:\Users\Acer\Desktop\Punjab EI\Punjab Transport.xlsx"
SRC_SHEET  = "Cleaned"
OUT_SHEET  = "Classified"

# Each PBS category maps to one or more (EI class, fraction) tuples.
CLASSIFICATION = {
    "Motor Cycles/Scooters":                          [("2W",  1.00)],
    "Auto Rickshaws":                                  [("3W",  1.00)],
    "Motor Cars/Jeeps/Station Wagons":                 [("4W1", 0.85),
                                                        ("4W2", 0.15)],   # 4W2 = diesel cars/SUVs only
    "Taxis":                                           [("4WT", 1.00)],
    "Pickups/Delivery Vans":                           [("LDV", 1.00)],   # Light Duty Vehicles
    "Mini Buses/Buses/Flying/Luxury Coaches":          [("BUS", 1.00)],
    "Trucks":                                          [("HDV", 1.00)],   # Heavy Duty Vehicles
    "Tractors":                                        [("NRV", 1.00)],
    "Other Vehicles":                                  [],                 # excluded
}

EI_CLASS_ORDER = ["2W", "3W", "4W1", "4W2", "4WT", "LDV", "HDV", "BUS", "NRV"]

df = pd.read_excel(INPUT_FILE, sheet_name=SRC_SHEET)
df.columns = ["Year", "Province", "Division", "District", "VehicleType", "Count"]
df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0)

print(f"Loaded {len(df)} rows from '{SRC_SHEET}'")

# Classify
records = []
for _, row in df.iterrows():
    vt = row["VehicleType"]
    mappings = CLASSIFICATION.get(vt, [])
    for ei_class, fraction in mappings:
        records.append({
            "Year":        int(row["Year"]),
            "Province":    row["Province"],
            "Division":    row["Division"],
            "District":    row["District"],
            "VehicleType": vt,
            "EI_Class":    ei_class,
            "Count":       round(row["Count"] * fraction),
        })

df_class = pd.DataFrame(records)
df_class = df_class.sort_values(["Year", "District", "EI_Class"]).reset_index(drop=True)

excluded = set(CLASSIFICATION.keys()) - {k for k, v in CLASSIFICATION.items() if v}
print(f"Excluded PBS categories: {excluded}")
print(f"Classified rows: {len(df_class)}")
print(f"\nClass totals (all years, all districts):")
totals = df_class.groupby("EI_Class")["Count"].sum()
for cls in EI_CLASS_ORDER:
    print(f"  {cls:6s}: {int(totals.get(cls, 0)):>15,}")

# Write to Excel (append mode — preserves all existing sheets)
with pd.ExcelWriter(INPUT_FILE, engine="openpyxl", mode="a",
                    if_sheet_exists="replace") as writer:
    df_class.to_excel(writer, sheet_name=OUT_SHEET, index=False)

    # Apply styling after the sheet has been written
    wb  = writer.book
    ws  = writer.sheets[OUT_SHEET]

    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    CLASS_COLORS = {
        "2W":  "E2EFDA",
        "3W":  "FFF2CC",
        "4W1": "DEEAF1",
        "4W2": "BDD7EE",
        "4WT": "FCE4D6",
        "LDV": "FFE699",
        "HDV": "EAD1DC",
        "BUS": "F4CCCC",
        "NRV": "D9D9D9",
    }
    col_widths = [6, 10, 14, 16, 42, 10, 14]

    # Style header row
    for c, (cell, w) in enumerate(zip(ws[1], col_widths), 1):
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 18

    # Colour-code data rows by EI_Class
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ei_class = row[5].value   # column F = EI_Class
        fill = PatternFill(start_color=CLASS_COLORS.get(ei_class, "FFFFFF"),
                           end_color=CLASS_COLORS.get(ei_class, "FFFFFF"),
                           fill_type="solid")
        for cell in row:
            cell.fill = fill
        row[6].number_format = "#,##0"   # Count column
        row[6].alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"

print(f"\nSaved '{OUT_SHEET}' sheet to {INPUT_FILE}")

# Verification pivot
print("\n=== CLASS TOTALS BY YEAR (Punjab-wide) ===")
pivot = df_class.pivot_table(index="Year", columns="EI_Class",
                              values="Count", aggfunc="sum")[EI_CLASS_ORDER]
pivot.columns.name = None
print(pivot.applymap(lambda x: f"{int(x):>12,}").to_string())
